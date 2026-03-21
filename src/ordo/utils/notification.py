import os
import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


# ---------------------------
# INDEX CALCULATION
# ---------------------------
def calculate_index(last_access_date, file_size_mb):
    today = datetime.datetime.now()
    days_unused = (today - last_access_date).days

    access_score = min(days_unused / 90, 1) * 100
    size_score = min(file_size_mb / 500, 1) * 100

    final_score = (0.7 * access_score) + (0.3 * size_score)
    return round(final_score, 2)


# ---------------------------
# MAIN ANALYSIS FUNCTION
# ---------------------------
def analyze_folder(folder_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "File Index Data"

    headers = ["File Index (%)", "File Name", "File Path", "Size (MB)", "Last Accessed"]
    ws.append(headers)

    # Bold header
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    risky_space = 0
    row_num = 2

    for root_dir, dirs, files in os.walk(folder_path):
        for file in files:
            file_path = os.path.join(root_dir, file)

            try:
                stat = os.stat(file_path)
                last_access = datetime.datetime.fromtimestamp(stat.st_atime)
                size_mb = stat.st_size / (1024 * 1024)

                index_value = calculate_index(last_access, size_mb)

                ws.append([
                    index_value,
                    file,
                    file_path,
                    round(size_mb, 2),
                    last_access.strftime("%Y-%m-%d")
                ])

                # Highlight risky files
                if index_value > 70:
                    for col in range(1, 6):
                        ws.cell(row=row_num, column=col).fill = PatternFill(
                            start_color="FF9999",
                            end_color="FF9999",
                            fill_type="solid"
                        )
                    risky_space += size_mb

                row_num += 1

            except Exception:
                continue

    # Auto column width
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    # Summary
    ws.append([])
    ws.append(["Total Risky Space (MB)", round(risky_space, 2)])

    wb.save("data.xlsx")

    messagebox.showinfo(
        "Analysis Complete",
        f"Excel file 'data.xlsx' created successfully.\n\nTotal Risky Space: {round(risky_space,2)} MB"
    )


# ---------------------------
# GUI + PERMISSION
# ---------------------------
def start_program():
    root = tk.Tk()
    root.withdraw()  # Hide main window

    permission = messagebox.askyesno(
        "Permission Required",
        "This tool will scan selected folder metadata only.\n\n"
        "No files will be modified or deleted.\n\n"
        "Do you allow access?"
    )

    if permission:
        folder_selected = filedialog.askdirectory(title="Select Folder to Analyze")

        if folder_selected:
            analyze_folder(folder_selected)
        else:
            messagebox.showwarning("No Folder", "No folder selected.")
    else:
        messagebox.showinfo("Permission Denied", "Operation cancelled by user.")


# ---------------------------
# RUN PROGRAM
# ---------------------------
if __name__ == "__main__":
    start_program()
    print("code run sucessfully.")